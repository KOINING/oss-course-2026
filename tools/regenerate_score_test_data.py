from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TEST_DATA_DIR = ROOT / "docs" / "test-data"
TEMPLATE_PATH = TEST_DATA_DIR / "template_actual.xlsx"


SCORES = {
    4: [12, 8, 18, 12, 8, 12, 18],
    5: [13, 9, 17, 13, 9, 13, 17],
    6: [11, 8, 16, 11, 8, 11, 16],
    7: [14, 9, 19, 14, 9, 14, 19],
    8: [10, 7, 15, 10, 7, 10, 15],
    9: [12, 8, 17, 12, 8, 12, 17],
    10: [9, 7, 14, 9, 7, 9, 14],
    11: [13, 8, 18, 13, 8, 13, 18],
    12: [12, 9, 17, 12, 9, 12, 17],
    13: [11, 8, 16, 11, 8, 11, 16],
    14: [14, 9, 19, 14, 9, 14, 19],
    15: [10, 7, 15, 10, 7, 10, 15],
    16: [12, 8, 17, 12, 8, 12, 17],
    17: [13, 9, 18, 13, 9, 13, 18],
    18: [11, 8, 16, 11, 8, 11, 16],
    19: [12, 8, 17, 12, 8, 12, 17],
    20: [14, 9, 19, 14, 9, 14, 19],
    21: [10, 7, 15, 10, 7, 10, 15],
    22: [13, 8, 18, 13, 8, 13, 18],
    23: [11, 8, 16, 11, 8, 11, 16],
    24: [12, 9, 17, 12, 9, 12, 17],
    25: [14, 9, 19, 14, 9, 14, 19],
}


def build_base_workbook():
    workbook = load_workbook(TEMPLATE_PATH)
    sheet = workbook[workbook.sheetnames[0]]
    for row_idx, scores in SCORES.items():
        for col_idx, score in enumerate(scores, start=3):
            sheet.cell(row_idx, col_idx).value = score
    return workbook, sheet


def save_valid():
    workbook, _ = build_base_workbook()
    workbook.save(TEST_DATA_DIR / "score_import_valid_class2.xlsx")


def save_invalid_rows():
    workbook, sheet = build_base_workbook()

    # 学号不属于当前教学班
    sheet.cell(5, 1).value = "99999999999"

    # 姓名与名单不一致
    sheet.cell(6, 2).value = "WrongName"

    # 成绩超出满分（第 1 个成绩列满分为 15）
    sheet.cell(7, 3).value = 18

    # 文件内重复学号
    sheet.cell(8, 1).value = sheet.cell(4, 1).value
    sheet.cell(8, 2).value = sheet.cell(4, 2).value

    workbook.save(TEST_DATA_DIR / "score_import_invalid_rows_class2.xlsx")


def save_invalid_header():
    workbook, sheet = build_base_workbook()
    sheet.cell(1, 3).value = "INVALID_HEADER"
    workbook.save(TEST_DATA_DIR / "score_import_invalid_header_class2.xlsx")


if __name__ == "__main__":
    save_valid()
    save_invalid_rows()
    save_invalid_header()
